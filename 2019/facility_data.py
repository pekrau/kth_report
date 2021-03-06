"""Facility Reports 2019.

Read the aggregated facility data XLSX files and return as lists of
dictionaries, one dictionary for each row.

The files were created in and downloaded from the Reporting Portal 
Reporting Portal https://reporting.scilifelab.se/
"""

import glob
import os.path
import unicodedata

import openpyxl

BASEDIRPATH = os.path.expanduser("~/Nextcloud/Årsrapport 2019/Facility reports")
BASEFILENAME = "orders_Facility_report_2019"

### Path to directory containing the downloaded aggregate files.
DIRPATH = os.path.join(BASEDIRPATH, "aggregate_files")

### Path to directory containing the downloaded volume data files.
VOLDIRPATH = os.path.join(BASEDIRPATH, "volume_data_files")


# Lookup from facility name to platform name.
PLATFORM_LOOKUP = {
    "Advanced Light Microscopy (ALM)" : "Cellular and Molecular Imaging",
    "Ancient DNA" : "Genomics",
    "Autoimmunity Profiling" : "Proteomics and Metabolomics",
    "BioImage Informatics" : "Cellular and Molecular Imaging",
    "Cell Profiling" : "Cellular and Molecular Imaging",
    "Chemical Biology Consortium Sweden" : "Chemical Biology and Genome Engineering",
    "Chemical Proteomics and Proteogenomics (MBB)" : "Proteomics and Metabolomics",
    "Chemical Proteomics and Proteogenomics (OnkPat)" : "Proteomics and Metabolomics",
    "Clinical Genomics Göteborg" : "Diagnostics Development",
    "Clinical Genomics Lund" : "Diagnostics Development",
    "Clinical Genomics Stockholm" : "Diagnostics Development",
    "Clinical Genomics Uppsala" : "Diagnostics Development",
    "Compute and Storage" : "Bioinformatics",
    "Cryo-EM (SU)" : "Cellular and Molecular Imaging",
    "Cryo-EM (UmU)" : "Cellular and Molecular Imaging",
    "Drug Discovery and Development" : "Drug Discovery and Development",
    # *** NOTE *** This is renamed to 'In Situ Sequencing', see below.
    # "Eukaryotic Single Cell Genomics" : "Genomics",
    "Genome Engineering Zebrafish" : "Chemical Biology and Genome Engineering",
    "High Throughput Genome Engineering" : "Chemical Biology and Genome Engineering",
    "In Situ Sequencing" : "Genomics",
    "Long-term Support (WABI)" : "Bioinformatics",
    "Mass Cytometry (KI)" : "Proteomics and Metabolomics",
    "Mass Cytometry (LiU)" : "Proteomics and Metabolomics",
    "Microbial Single Cell Genomics" : "Genomics",
    "NGI Stockholm" : "Genomics",
    "NGI Uppsala SNP&SEQ" : "Genomics",
    "NGI Uppsala UGC" : "Genomics",
    "PLA and Single Cell Proteomics" : "Proteomics and Metabolomics",
    "Plasma Profiling" : "Proteomics and Metabolomics",
    "Protein Science Facility" : "Cellular and Molecular Imaging",
    "Support and Infrastructure" : "Bioinformatics",
    "Swedish Metabolomics Centre" : "Proteomics and Metabolomics",
    "Swedish NMR Centre" : "Cellular and Molecular Imaging",
    "Systems Biology" : "Bioinformatics"
}

# Handle strange cases where facility name has wrong character case.
PLATFORM_LOOKUP_LOWER = dict([(k.lower(), (k, v))
                              for k, v in PLATFORM_LOOKUP.items()])

REPORT_FILEPATH = os.path.join(DIRPATH, f"{BASEFILENAME}.xlsx")

def get_report_data(filepath=REPORT_FILEPATH):
    """Get the data reported in single-valued fields of the
    OrderPortal form.
    """
    return read_file(filepath)

FACILITY_HEAD_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_head.xlsx")

def get_facility_head_data(filepath=FACILITY_HEAD_FILEPATH):
    """Get the facility head data.
    """
    return read_file(filepath)

FACILITY_DIRECTOR_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_director.xlsx")

def get_facility_director_data(filepath=FACILITY_DIRECTOR_FILEPATH):
    """Get the facility director data.
    """
    return read_file(filepath)

ADDITIONAL_FUNDING_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_additional_funding.xlsx")

def get_additional_funding_data(filepath=ADDITIONAL_FUNDING_FILEPATH):
    """Get the additional funding data.
    """
    return read_file(filepath)

IP_RIGHTS_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_immaterial_property_rights.xlsx")

def get_ip_rights_data(filepath=IP_RIGHTS_FILEPATH):
    """Get the immaterial_propery_rights data."""
    return read_file(filepath)

def read_file(filepath):
    """Open the Excel file given by the path and read the first sheet.
    Return a list of dictionaries, one for each row.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    rows = list(wb.active)
    header = [c.value.strip() for c in rows[0]]
    result = []
    for row in rows[1:]:
        record = dict(list(zip(header, [c.value for c in row])))
        # *** NOTE *** Special case; correct the name of one facility.
        if record.get("facility") == "Eukaryotic Single Cell Genomics":
            record["facility"] = "In Situ Sequencing"
        result.append(record)
    return result

def get_volume_data(sheetname, dirpath=VOLDIRPATH):
    """Get all data records for a specified sheet for each facility.
    Returns list of dictionaries, where each dictionary is one row.
    """
    skip_rows_until = "Name of reporting unit"
    result = []
    for filepath in sorted(glob.glob(f"{dirpath}/*.xls[mx]")):
        try:
            records = read_volume_file(filepath, sheetname, skip_rows_until)
            result.extend(records)
            print(os.path.basename(filepath), len(records))
        except KeyError as error:
            print(os.path.basename(filepath), error)
            raise
    return result

def read_volume_file(filepath, sheetname, skip_rows_until):
    """Open the Excel Volume data file given by the path and read the
    sheet with the given name, or the first sheet. Return a list
    of records, where a record is a dictionary with keys from the
    header row in the sheet.
    NOTE: Special case; correct the name of one facility.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.get_sheet_by_name(sheetname)

    # Awful kludge to avoid weird behaviour when reading
    # all rows from a sheet in one file after another...
    # Get chunks of rows until the first cell has no value.
    rows = []
    while True:
        for row in ws.iter_rows(min_row=len(rows)+1, max_row=len(rows)+100):
            # Yet another kludge to handle special case where a field
            # supposed to contain an email address instead contains
            # a formula that computes the email address from the name
            # and a given domain name. I am impressed, kind of...
            # This relies on the formula getting the first name and last
            # name from the two preceding columns, and having the '@'
            # before the domain name.
            values = []
            for pos, cell in enumerate(row):
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    value = row[pos-2].value 
                    value += "." + row[pos-1].value
                    value += cell.value[cell.value.index("@"):].rstrip('"')
                    value = to_ascii(value)
                    value = value.replace(" ", "-")
                values.append(value)
            rows.append(values)
        if not rows[-1] or rows[-1][0] is None: break
    # Remove empty rows at the end of the list.
    while rows[-1][0] is None:
        rows.pop()

    wb.close()

    # Find the header row.
    for first, row in enumerate(rows):
        if row[0] and skip_rows_until in row[0]: break
    headers = [c.strip() for c in rows[first] if c is not None]

    # Find the key for the facility name. Groan! This is just terrible...
    # The reason is that the header "1. Name of reporting unit..."
    # sometimes has one white-space after "1.", sometimes two.
    for key in rows[first]:
        if "Name of reporting unit" in key: break
    else:
        raise KeyError("Sorry, could not find the facility name column.")

    result = []
    for row in rows[first+1:]:
        record = dict(zip(headers, row))
        # *** NOTE *** Special case; correct the name of one facility.
        if record.get(key) == "Eukaryotic Single Cell Genomics":
            record[key] = "In Situ Sequencing"
        result.append(record)

    # Modify the key for the facility name to contain only single white-space.
    proper_key = " ".join(key.split())
    if proper_key != key:
        for record in result:
            record[proper_key] = record.pop(key)

    return result

def to_ascii(value):
    "Convert any non-ASCII character to its closest ASCII equivalent."
    if value is None: return ''
    value = unicodedata.normalize('NFKD', str(value))
    return u''.join([c for c in value if not unicodedata.combining(c)])


if __name__ == "__main__":
    print(len(get_users_data()))
