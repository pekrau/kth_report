"""Unit Reports 2021.

[Keep the old name 'facility_data.py' even though the term is now 'unit'.

Read the aggregated unit data XLSX files and return as lists of
dictionaries, one dictionary for each row.

The files were created in and downloaded from the Reporting Portal 
Reporting Portal https://reporting.scilifelab.se/
"""

import glob
import os.path
import unicodedata

import openpyxl

BASEDIRPATH = os.path.expanduser("~/Nextcloud/Årsrapport 2021/Unit reports")
BASEFILENAME = "orders_Units_report_2021"

### Path to directory containing the downloaded aggregate files.
DIRPATH = os.path.join(BASEDIRPATH, "aggregate_files")

### Path to directory containing the downloaded volume data files.
VOLDIRPATH = os.path.join(BASEDIRPATH, "volume_data_files")


# Lookup from unit name to platform name.
# Information from the file "Reporting Units 2021.xlsx"
PLATFORM_LOOKUP = {
    "AIDA Data Hub": "Bioinformatics",
    "Compute and Storage" : "Bioinformatics",
    "BioImage Informatics" : "Bioinformatics",
    "Support, Infrastructure and Training" : "Bioinformatics",

    "Ancient DNA" : "Genomics",
    "Microbial Single Cell Genomics" : "Genomics",
    "National Genomics Infrastructure" : "Genomics",

    "Clinical Genomics Gothenburg" : "Clinical Genomics",
    "Clinical Genomics Linköping" : "Clinical Genomics",
    "Clinical Genomics Lund" : "Clinical Genomics",
    "Clinical Genomics Stockholm" : "Clinical Genomics",
    "Clinical Genomics Umeå" : "Clinical Genomics",
    "Clinical Genomics Uppsala" : "Clinical Genomics",
    "Clinical Genomics Örebro" : "Clinical Genomics",

    "Autoimmunity and Serology Profiling" : "Clinical Proteomics and Immunology",
    "Affinity Proteomics Stockholm" : "Clinical Proteomics and Immunology",
    "Affinity Proteomics Uppsala" : "Clinical Proteomics and Immunology",
    "Cellular Immunomonitoring" : "Clinical Proteomics and Immunology",
    "Global Proteomics and Proteogenomics" : "Clinical Proteomics and Immunology",
    "Glycoproteomics" : "Clinical Proteomics and Immunology",

    "Swedish Metabolomics Centre" : "Metabolomics",
    "Exposomics" : "Metabolomics",

    "Eukaryotic Single Cell Genomics" : "Spatial and Single Cell Biology",
    "Spatial Proteomics" : "Spatial and Single Cell Biology",
    "In Situ Sequencing" : "Spatial and Single Cell Biology",
    "Advanced FISH Technologies" : "Spatial and Single Cell Biology",
    "Spatial Mass Spectrometry" : "Spatial and Single Cell Biology",

    "Cryo-EM" : "Cellular and Molecular Imaging",
    "Integrated Microscopy Technologies Gothenburg" : "Cellular and Molecular Imaging",
    "Integrated Microscopy Technologies Stockholm" : "Cellular and Molecular Imaging",
    "Integrated Microscopy Technologies Umeå" : "Cellular and Molecular Imaging",

    "Swedish NMR Centre" : "Integrated Structural Biology",
    "Structural Proteomics" : "Integrated Structural Biology",

    "Chemical Biology Consortium Sweden" : "Chemical Biology and Genome Engineering",
    "Chemical Proteomics" : "Chemical Biology and Genome Engineering",
    "CRISPR Functional Genomics" : "Chemical Biology and Genome Engineering",
    "Genome Engineering Zebrafish" : "Chemical Biology and Genome Engineering",

    "Drug Discovery and Development" : "Drug Discovery and Development",
}

# Handle strange cases where unit name has wrong character case.
PLATFORM_LOOKUP_LOWER = dict([(k.lower(), (k, v))
                              for k, v in PLATFORM_LOOKUP.items()])

REPORT_FILEPATH = os.path.join(DIRPATH, f"{BASEFILENAME}.xlsx")

def get_report_data(filepath=REPORT_FILEPATH):
    """Get the data reported in single-valued fields of the
    OrderPortal form.
    """
    return read_file(filepath)

FACILITY_HEAD_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_head.xlsx")

def get_facility_head_data(filepath=FACILITY_HEAD_FILEPATH):
    """Get the facility head data.
    """
    return read_file(filepath)

FACILITY_DIRECTOR_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_director.xlsx")

def get_facility_director_data(filepath=FACILITY_DIRECTOR_FILEPATH):
    """Get the facility director data.
    """
    return read_file(filepath)

ADDITIONAL_FUNDING_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_additional_funding.xlsx")

def get_additional_funding_data(filepath=ADDITIONAL_FUNDING_FILEPATH):
    """Get the additional funding data.
    """
    return read_file(filepath)

IP_RIGHTS_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_immaterial_property_rights.xlsx")

def get_ip_rights_data(filepath=IP_RIGHTS_FILEPATH):
    """Get the immaterial_propery_rights data."""
    return read_file(filepath)

def read_file(filepath):
    """Open the Excel file given by the path and read the first sheet.
    Return a list of dictionaries, one for each row.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    rows = list(wb.active)
    wb.close()
    header = [c.value.strip() for c in rows[0]]
    return [dict(list(zip(header, [c.value for c in row])))
            for row in rows[1:]]

def get_volume_data(sheetname, dirpath=VOLDIRPATH):
    """Get all data records for a specified sheet for each unit.
    Returns list of dictionaries, where each dictionary is one row.
    """
    skip_rows_until = "Name of reporting unit"
    result = []
    for filepath in sorted(glob.glob(f"{dirpath}/*.xls[mx]")):
        try:
            records = read_volume_file(filepath, sheetname, skip_rows_until)
            result.extend(records)
            print(os.path.basename(filepath), len(records))
        except KeyError as error:
            print(os.path.basename(filepath), error)
            raise
    return result

def read_volume_file(filepath, sheetname, skip_rows_until):
    """Open the Excel Volume data file given by the path and read the
    sheet with the given name, or the first sheet. Return a list
    of records, where a record is a dictionary with keys from the
    header row in the sheet.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.get_sheet_by_name(sheetname)

    # Awful kludge to avoid weird behaviour when reading
    # all rows from a sheet in one file after another...
    # Get chunks of rows until the first cell has no value.
    rows = []
    while True:
        for row in ws.iter_rows(min_row=len(rows)+1, max_row=len(rows)+100):
            # Yet another kludge to handle special case where a field
            # supposed to contain an email address instead contains
            # a formula that computes the email address from the name
            # and a given domain name. I am impressed, kind of...
            # This relies on the formula getting the first name and last
            # name from the two preceding columns, and having the '@'
            # before the domain name.
            values = []
            for pos, cell in enumerate(row):
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    value = row[pos-2].value 
                    value += "." + row[pos-1].value
                    value += cell.value[cell.value.index("@"):].rstrip('"')
                    value = to_ascii(value)
                    value = value.replace(" ", "-")
                values.append(value)
            rows.append(values)
        if not rows[-1] or rows[-1][0] is None: break
    # Remove empty rows at the end of the list.
    while rows[-1][0] is None:
        rows.pop()

    wb.close()

    # Find the header row.
    for first, row in enumerate(rows):
        if row[0] and skip_rows_until in row[0]: break
    headers = [c.strip() for c in rows[first] if c is not None]

    # Find the key for the facility name. Groan! This is just terrible...
    # The reason is that the header "1. Name of reporting unit..."
    # sometimes has one white-space after "1.", sometimes two.
    for key in rows[first]:
        if "Name of reporting unit" in key: break
    else:
        raise KeyError("Sorry, could not find the facility name column.")

    result = [dict(zip(headers, row)) for row in rows[first+1:]]

    # Modify the key for the unit name to contain only single white-space.
    proper_key = " ".join(key.split())
    if proper_key != key:
        for record in result:
            record[proper_key] = record.pop(key)

    return result


if __name__ == "__main__":
    units = sorted(PLATFORM_LOOKUP.keys(), key=lambda k: k.lower())
    for unit in units:
        print(unit)

    print()
    platforms = sorted(set(PLATFORM_LOOKUP.values()))
    for platform in platforms:
        print(platform)
