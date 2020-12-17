"""Merge the Volume Data files for Fellows reporting.
Per Kraulis 2020-09-23, 2020-12-03
"""

import fnmatch
import os
import os.path
import sys

import openpyxl
import xlsxwriter

INPUT_DIRPATH = os.path.expanduser('~/Nextcloud/Årsrapport 2020/Fellows reports/volume_data_files')
OUTPUT_DIRPATH = os.path.expanduser('~/Nextcloud/Årsrapport 2020/Fellows reports/merged_files')

teaching = []
teaching_header = None
TEACHING = '2. Teaching'
TEACHING_HEADER = '1. Name of activity*'
TEACHING_COLS = 11
TEACHING_DATE_COLS = (7, 8)     # After swapping name first
TEACHING_OUTPUT_NAME = 'Fellows 2020 Teaching.xlsx'

grants = []
grants_header = None
GRANTS = '4. Grants'
GRANTS_HEADER = '1. Name of grant*'
GRANTS_COLS = 7
GRANTS_OUTPUT_NAME = 'Fellows 2020 Grants.xlsx'

collaborations = []
collaborations_header = None
COLLABORATIONS = '5. Collaborations'
COLLABORATIONS_HEADER = '1. Name of organization*'
COLLABORATIONS_COLS = 8
COLLABORATIONS_OUTPUT_NAME = 'Fellows 2020 Collaborations.xlsx'

count = 0
for filename in os.listdir(INPUT_DIRPATH):
    if not fnmatch.fnmatch(filename, "*.xls[xm]"): continue
    print(filename)
    count += 1
    filepath = os.path.join(INPUT_DIRPATH, filename)
    wb =  openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Teaching data.
    try:
        ws = wb.get_sheet_by_name(TEACHING)
    except KeyError as msg:
        sys.exit(f"{filename}: {msg}")
    rows = list(ws.values)
    # Find the header row of the data.
    for pos, row in enumerate(rows):
        if row[0] == TEACHING_HEADER:
            teaching_header = list(row[:TEACHING_COLS])
            # Swap email and name columns to beginning.
            teaching_header.insert(0, teaching_header.pop())
            teaching_header.insert(0, teaching_header.pop())
            # Fix up cells that contained macros: not properly read.
            teaching_header[3] = '2. Level of education*'
            teaching_header[4] = '3. Type of activity?*'
            teaching_header[5] = '5. Did you have main responsibility of this educational activity?*'
            break
    else:
        sys.exit(f"{filename}: Could not find data rows for '{TEACHING}'")
    for row in rows[pos+1:]:
        if row[0] is None: continue
        row = list(row[:TEACHING_COLS])
        # Swap email and name columns to beginning.
        row.insert(0, row.pop())
        row.insert(0, row.pop())
        for date_col in TEACHING_DATE_COLS:
            if row[date_col]:
                row[date_col] = row[date_col].strftime('%Y-%m-%d')
        teaching.append(row)

    # Grants data.
    try:
        ws = wb.get_sheet_by_name(GRANTS)
    except KeyError as msg:
        sys.exit(f"{filename}: {msg}")
    rows = list(ws.values)
    # Find the header row of the data.
    for pos, row in enumerate(rows):
        if row[0] == GRANTS_HEADER:
            grants_header = list(row[:GRANTS_COLS])
            # Swap email and name columns to beginning.
            grants_header.insert(0, grants_header.pop())
            grants_header.insert(0, grants_header.pop())
            # Fix up cells that contained macros: not properly read.
            grants_header[4] = '3. Type of grant*'
            break
    else:
        sys.exit(f"{filename}: Could not find data rows for '{GRANTS}'")
    for row in rows[pos+1:]:
        if row[0] is None: continue
        row = list(row[:GRANTS_COLS])
        # Swap email and name columns to beginning.
        row.insert(0, row.pop())
        row.insert(0, row.pop())
        grants.append(row)

    # Collaborations data.
    try:
        ws = wb.get_sheet_by_name(COLLABORATIONS)
    except KeyError as msg:
        sys.exit(f"{filename}: {msg}")
    rows = list(ws.values)
    # Find the header row of the data.
    for pos, row in enumerate(rows):
        if row[0] == COLLABORATIONS_HEADER:
            collaborations_header = list(row[:COLLABORATIONS_COLS])
            # Swap email and name columns to beginning.
            collaborations_header.insert(0, collaborations_header.pop())
            collaborations_header.insert(0, collaborations_header.pop())
            # Fix up cells that contained macros: not properly read.
            collaborations_header[3] = '2. Type of organization*'
            collaborations_header[6] = '5. Collaboration formed during 2020?*'
            break
    else:
        sys.exit(f"{filename}: Could not find data rows for '{COLLABORATIONS}'")
    for row in rows[pos+1:]:
        if row[0] is None: continue
        row = list(row[:COLLABORATIONS_COLS])
        # Swap email and name columns to beginning.
        row.insert(0, row.pop())
        row.insert(0, row.pop())
        collaborations.append(row)

print("Read", count, "input Volume data files.")

# Output merged Teaching data.
wb = xlsxwriter.Workbook(os.path.join(OUTPUT_DIRPATH, TEACHING_OUTPUT_NAME))
head_text_format = wb.add_format({'bold':True,
                                  'text_wrap':True,
                                  'bg_color':'#9ECA7F',
                                  'font_size':15,
                                  'align':'center',
                                  'border':1})
normal_text_format = wb.add_format({'font_size':14,
                                    'align':'left',
                                    'valign':'vcenter'})
long_text_format = wb.add_format({'text_wrap':True,
                                  'font_size':14,
                                  'align':'left',
                                  'valign':'vcenter'})
ws = wb.add_worksheet()
ws.freeze_panes(1, 2)
ws.set_row(0, None, head_text_format)
ws.set_column(0, 1, 20, normal_text_format)
ws.set_column(2, 2, 20, normal_text_format)
ws.set_column(3, 5, 10, normal_text_format)
ws.set_column(6, 6, 5, normal_text_format)
ws.set_column(7, 8, 12, normal_text_format)
ws.set_column(9, 9, 5, normal_text_format)
ws.set_column(10, 10, 60, long_text_format)

ws.write_row(0, 0, teaching_header)
for pos, row in enumerate(teaching, 1):
    ws.write_row(pos, 0, row)
wb.close()

# Output merged Grants data.
wb = xlsxwriter.Workbook(os.path.join(OUTPUT_DIRPATH, GRANTS_OUTPUT_NAME))
head_text_format = wb.add_format({'bold':True,
                                  'text_wrap':True,
                                  'bg_color':'#9ECA7F',
                                  'font_size':15,
                                  'align':'center',
                                  'border':1})
normal_text_format = wb.add_format({'font_size':14,
                                    'align':'left',
                                    'valign':'vcenter'})
long_text_format = wb.add_format({'text_wrap':True,
                                  'font_size':14,
                                  'align':'left',
                                  'valign':'vcenter'})
ws = wb.add_worksheet()
ws.freeze_panes(1, 2)
ws.set_row(0, None, head_text_format)
ws.set_column(0, 1, 20, normal_text_format)
ws.set_column(2, 2, 40, normal_text_format)
ws.set_column(3, 3, 30, normal_text_format)
ws.set_column(4, 4, 16, normal_text_format)
ws.set_column(5, 5, 10, normal_text_format)
ws.set_column(6, 6, 60, long_text_format)

ws.write_row(0, 0, grants_header)
for pos, row in enumerate(grants, 1):
    ws.write_row(pos, 0, row)
wb.close()

# Output merged Collaborations data.
wb = xlsxwriter.Workbook(os.path.join(OUTPUT_DIRPATH, COLLABORATIONS_OUTPUT_NAME))
head_text_format = wb.add_format({'bold':True,
                                  'text_wrap':True,
                                  'bg_color':'#9ECA7F',
                                  'font_size':15,
                                  'align':'center',
                                  'border':1})
normal_text_format = wb.add_format({'font_size':14,
                                    'align':'left',
                                    'valign':'vcenter'})
long_text_format = wb.add_format({'text_wrap':True,
                                  'font_size':14,
                                  'align':'left',
                                  'valign':'vcenter'})
ws = wb.add_worksheet()
ws.freeze_panes(1, 2)
ws.set_row(0, None, head_text_format)
ws.set_column(0, 1, 20, normal_text_format)
ws.set_column(2, 2, 40, normal_text_format)
ws.set_column(3, 3, 30, normal_text_format)
ws.set_column(4, 4, 16, normal_text_format)
ws.set_column(5, 5, 20, normal_text_format)
ws.set_column(6, 6, 5, normal_text_format)
ws.set_column(7, 7, 60, long_text_format)

ws.write_row(0, 0, collaborations_header)
for pos, row in enumerate(collaborations, 1):
    ws.write_row(pos, 0, row)
wb.close()
