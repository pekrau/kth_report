"""Facility Reports 2020.

Read the aggregated facility data XLSX files and return as lists of
dictionaries, one dictionary for each row.

The files were created in and downloaded from the Reporting Portal 
Reporting Portal https://reporting.scilifelab.se/
"""

import glob
import os.path

import openpyxl

BASEDIRPATH = os.path.expanduser("~/Nextcloud/Årsrapport 2020/Facility reports")
BASEFILENAME = "orders_Facility_report_2020"

### Path to directory containing the downloaded aggregate files.
DIRPATH = os.path.join(BASEDIRPATH, "aggregate_files")

### Path to directory containing the downloaded volume data files.
VOLDIRPATH = os.path.join(BASEDIRPATH, "volume_data_files")


# Lookup from facility name to platform name.
# Information from the file "Reporting Units 2020.xlsx"
PLATFORM_LOOKUP = {
    "Advanced Light Microscopy" : "Cellular and Molecular Imaging",
    "Ancient DNA" : "Genomics",
    "Autoimmunity and Serology Profiling" : "Proteomics and Metabolomics",
    "BioImage Informatics" : "Cellular and Molecular Imaging",
    "Cell Profiling" : "Cellular and Molecular Imaging",
    "Chemical Biology Consortium Sweden" : "Chemical Biology and Genome Engineering",
    "Chemical Proteomics and Proteogenomics (MBB)" : "Proteomics and Metabolomics",
    "Chemical Proteomics and Proteogenomics (OnkPat)" : "Proteomics and Metabolomics",
    "Clinical Genomics Gothenburg" : "Diagnostics Development",
    "Clinical Genomics Linköping" : "Diagnostics Development",
    "Clinical Genomics Lund" : "Diagnostics Development",
    "Clinical Genomics Stockholm" : "Diagnostics Development",
    "Clinical Genomics Umeå" : "Diagnostics Development",
    "Clinical Genomics Uppsala" : "Diagnostics Development",
    "Clinical Genomics Örebro" : "Diagnostics Development",
    "Compute and Storage" : "Bioinformatics",
    "Cryo-EM" : "Cellular and Molecular Imaging",
    "Drug Discovery and Development" : "Drug Discovery and Development",
    "Eukaryotic Single Cell Genomics" : "Genomics",
    "Genome Engineering Zebrafish" : "Chemical Biology and Genome Engineering",
    "High Throughput Genome Engineering" : "Chemical Biology and Genome Engineering",
    "Long-term Support (WABI)" : "Bioinformatics",
    "Mass Cytometry (KI)" : "Proteomics and Metabolomics",
    "Mass Cytometry (LiU)" : "Proteomics and Metabolomics",
    "Microbial Single Cell Genomics" : "Genomics",
    "National Genomics Infrastructure" : "Genomics",
    "PLA and Single Cell Proteomics" : "Proteomics and Metabolomics",
    "Plasma Profiling" : "Proteomics and Metabolomics",
    "Protein Science Facility" : "Cellular and Molecular Imaging",
    "Support and Infrastructure" : "Bioinformatics",
    "Swedish Metabolomics Centre" : "Proteomics and Metabolomics",
    "Swedish NMR Centre" : "Cellular and Molecular Imaging",
    "Systems Biology" : "Bioinformatics"
}

# Handle strange cases where facility name has wrong character case.
PLATFORM_LOOKUP_LOWER = dict([(k.lower(), (k, v))
                              for k, v in PLATFORM_LOOKUP.items()])

REPORT_FILEPATH = os.path.join(DIRPATH, f"{BASEFILENAME}.xlsx")

def get_report_data(filepath=REPORT_FILEPATH):
    """Get the data reported in single-valued fields of the
    OrderPortal form.
    """
    return read_file(filepath)

FACILITY_HEAD_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_head.xlsx")

def get_facility_head_data(filepath=FACILITY_HEAD_FILEPATH):
    """Get the facility head data.
    """
    return read_file(filepath)

FACILITY_DIRECTOR_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_facility_director.xlsx")

def get_facility_director_data(filepath=FACILITY_DIRECTOR_FILEPATH):
    """Get the facility director data.
    """
    return read_file(filepath)

ADDITIONAL_FUNDING_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_additional_funding.xlsx")

def get_additional_funding_data(filepath=ADDITIONAL_FUNDING_FILEPATH):
    """Get the additional funding data.
    """
    return read_file(filepath)

IP_RIGHTS_FILEPATH = os.path.join(
    DIRPATH, f"{BASEFILENAME}_immaterial_property_rights.xlsx")

def get_ip_rights_data(filepath=IP_RIGHTS_FILEPATH):
    """Get the immaterial_propery_rights data."""
    return read_file(filepath)

def read_file(filepath):
    """Open the Excel file given by the path and read the first sheet.
    Return a list of dictionaries, one for each row.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    rows = list(wb.active)
    wb.close()
    header = [c.value.strip() for c in rows[0]]
    return [dict(list(zip(header, [c.value for c in row])))
            for row in rows[1:]]

def get_users_data(dirpath=VOLDIRPATH):
    """Get all users for each facility.
    Returns list of dictionaries, where each dictionary is one row.
    """
    count = 0
    title = "1.  Name of reporting unit* (choose from drop-down menu)"
    result = []
    for filepath in sorted(glob.glob(f"{dirpath}/*.xls[mx]")):
        try:
            records = read_volume_file(filepath, "A. Users", title)
            result.extend(records)
            print(os.path.basename(filepath), len(records))
        except KeyError as error:
            print(os.path.basename(filepath), error)
    return result

def read_volume_file(filepath, sheetname, skip_rows_until):
    """Open the Excel Volume data file given by the path and read the
    sheet with the given name, or the first sheet. Return a list
    of records, where a record is a dictionary with keys from the
    header row in the sheet.
    NOTE: Special case; correct the name of one facility.
    """
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.get_sheet_by_name(sheetname)

    # Awful kludge to avoid weird behaviour when reading
    # all rows from a sheet in one file after another...
    # Get chunks of rows until the first cell has no value.
    rows = []
    while True:
        for row in ws.iter_rows(min_row=len(rows)+1, max_row=len(rows)+100):
            rows.append([cell.value for cell in row])
        if not rows[-1] or rows[-1][0] is None: break
    # Remove empty rows at the end of the list.
    while rows[-1][0] is None:
        rows.pop()

    wb.close()

    # Find the header row.
    for first, row in enumerate(rows):
        if row[0] and skip_rows_until == row[0]: break
    headers = [c.strip() for c in rows[first] if c is not None]

    # result = []
    # for row in rows[first+1:]:
    #     result.append(dict(zip(headers, row)))

    return [dict(zip(headers, row)) for row in rows[first+1:]]


if __name__ == "__main__":
    print(len(get_users_data()))
